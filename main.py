# from PyQt6.QtWidgets import QMainWindow, QApplication, QPushButton, QFileDialog
from PyQt6.QtWidgets import *
from PyQt6.QtCore import Qt
import os
import json
import openpyxl
import time
import sys


def path_builder(path, file_name):
    if '\\' in path:
        return f"{path}\{file_name}"
    else:
        return f"{path}/{file_name}"


def not_empty_list(lst):
    return max([_ is not None for _ in lst])


class Main(QMainWindow):
    xlsx_addr = ''
    folder_addr = ''
    delete_flag = False

    def __init__(self):
        super().__init__()

        file_selection_text = QLabel(self)
        file_selection_text.setText("Select xlsx file: ")
        file_selection = QPushButton(self)
        file_selection.setText("...")
        file_selection.clicked.connect(lambda: self.open_dialog())

        folder_selection_text = QLabel(self)
        folder_selection_text.setText("Select folder for jsons: ")
        # folder_selection_value_text = QLabel(self)
        # folder_selection_value_text.setText(f'{self.folder_addr}')
        folder_selection = QPushButton(self)
        folder_selection.setText("...")
        folder_selection.clicked.connect(lambda: self.select_folder())

        checkbox_text = QLabel(self)
        checkbox_text.setText("Check this field to delete older jsons")
        checkbox = QCheckBox()
        checkbox.setCheckState(Qt.CheckState.Unchecked)
        checkbox.stateChanged.connect(self.show_state)

        run_button = QPushButton(self)
        run_button.setText("Run")
        run_button.clicked.connect(lambda: self.main_function())
        # run_button.clicked()

        layout = QGridLayout()
        layout.addWidget(file_selection_text, 0, 0)
        layout.addWidget(file_selection, 0, 1)
        layout.addWidget(folder_selection_text, 1, 0)
        layout.addWidget(folder_selection, 1, 1)
        layout.addWidget(checkbox_text, 2, 0)
        layout.addWidget(checkbox, 2, 1)
        layout.addWidget(run_button, 3, 0)

        widget = QWidget()
        widget.setLayout(layout)
        self.setCentralWidget(widget)

    @classmethod
    def show_state(cls, s):
        cls.deleteFlag = s

    @classmethod
    def open_dialog(cls):
        fname = QFileDialog.getOpenFileName(
            QWidget(),
            "Open File",
            "${HOME}",
            "All Files (*);; Python Files (*.py);; PNG Files (*.png)",
        )
        cls.xlsx_addr = fname[0]
        # print(cls.xlsx_addr)

    @classmethod
    def select_folder(cls):
        dir_name = QFileDialog.getExistingDirectory(
            QWidget(),
            "Select folder"
        )
        cls.folder_addr = dir_name
        # print(cls.folder_addr)

    @classmethod
    def main_function(cls):
        if cls.delete_flag:
            print("ну и хуй с тобой")
        else:
            cls.delete_old_files()

        # print(cls.xlsx_addr)
        excel_data = cls.read_excel()
        for row in excel_data:
            if row[2] is not None:
                cls.create_json(row[2], row[3])

    @classmethod
    def delete_old_files(cls):
        now = time.time()

        files = [os.path.join(cls.folder_addr, filename) for filename in os.listdir(cls.folder_addr)]
        for filename in files:
            if ".json" in filename:
                if (now - os.stat(filename).st_mtime) > 100:
                    os.remove(filename)

    @classmethod
    def create_json(cls, log, passw):
        data_dict = {
            "Enabled": True,
            "Paused": True,
            "SteamLogin": f"{log}",
            "SteamPassword": f"{passw}",
        }
        with open(path_builder(cls.folder_addr, log) + '.json', "w+") as file:
            json.dump(data_dict, file)

    @classmethod
    def read_excel(cls):
        xlsx_data = openpyxl.load_workbook(cls.xlsx_addr)
        sheet_data = xlsx_data.active
        max_row_count = sheet_data.max_row
        max_col_count = sheet_data.max_column
        rows_list = []
        for row in range(1, max_row_count + 1):
            row_data = []
            for col in range(1, max_col_count + 1):
                row_data.append(sheet_data.cell(row=row, column=col).value)
            if not_empty_list(row_data):
                rows_list.append(row_data)
        return rows_list


if __name__ == "__main__":
    app = QApplication(sys.argv)
    main_gui = Main()
    main_gui.show()
    # print(Main.folder_addr)
    # print(Main.xlsx_addr)
    sys.exit(app.exec())
